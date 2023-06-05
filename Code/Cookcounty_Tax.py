from datetime import datetime
import os, shutil
from selenium.common import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import pandas as pd
import shutil, os
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
import openpyxl
import Code.New_update1_title
import Code.Lien_Report
import Code.BRB_Search
import getOrders


def createZipfile(orderId):
    par_dir = os.getcwd()
    # path to folder which needs to be zipped
    directory = par_dir + '\\Output\\COOK_COUNTY\\Order No ' + str(orderId)
    #zip the order output folder
    shutil.make_archive(par_dir + '\\Output\\COOK_COUNTY\\Order', 'zip', directory)
    print('All files zipped successfully!')

def Final_UI(file):

 par_dir=os.getcwd()
 time.sleep(1)

 dataframe1 = pd.read_excel(par_dir+'\\Input\\'+file,engine='openpyxl')


 E = dataframe1[dataframe1.columns[0]].count()

 for i in range(E):
     try:
         workbook = openpyxl.load_workbook(os.getcwd() + '\\Input\\' + file)
         worksheet = workbook.active
         start_time = datetime.now()

         worksheet['j' + str(int(i + 2))] = start_time

         workbook.save(os.getcwd() + '\\Input\\' + file)

         EXCELADDRESS = str(dataframe1['Property Address'][i].replace("-", '')).lower()

         FName = (dataframe1['NAME'][i])
         F = FName.split()[0:1]
         L = FName.split()[-1]

         HOUSENUMBER = (EXCELADDRESS.split()[0:1])
         STREETNAME = (EXCELADDRESS.split()[1:3])

         STREETNAME = (" ".join(STREETNAME)).lower()

         if "avenue" in STREETNAME:  # added to replace AVenue, AVENUE words and also to remove after the Avenue word
             STREETNAME = STREETNAME.replace("avenue", "ave")
             STREETNAME = STREETNAME.split("ave")
             STREETNAME = STREETNAME[0]

         elif "ave" in STREETNAME:  # added to replace AVenue, AVENUE words and also to remove after the Avenue word

             STREETNAME = STREETNAME.split("ave")
             STREETNAME = STREETNAME[0]

         # print(STREETNAME)

         OrderID = int(dataframe1['Order ID'][i])
         OrderNum = dataframe1['Order No'][i]
         processId = int(dataframe1['Process ID'][i])

         getOrders.updateStatus(OrderID,OrderNum,"In Progress",processId,"Automation started")
         City = str(dataframe1['City'][i])

         # CC=City.split()[-1]

         PIN = str(dataframe1['Zip'][i])
         # PIN=PIN.split("-")[-1]

         chrome_options = Options()
         chrome_options.add_argument('--kiosk-printing')
         chrome_options.add_argument('--disable-extensions')

         driver = webdriver.Chrome(options=chrome_options)
         time.sleep(4)
         driver.maximize_window()
         driver.get('https://www.cookcountytreasurer.com/setsearchparameters.aspx')  # opening Tax page
         # driver.maximize_window()
         print("Running Tax Page")
         elem = WebDriverWait(driver, 15).until(
             EC.presence_of_element_located(
                 (By.XPATH, "/html/body/form/div[4]/div[2]/div/div/div[2]/div/div/ul/li[3]/div/span")))
         time.sleep(2)
         driver.find_element(By.XPATH, '/html/body/form/div[4]/div[2]/div/div/div[1]/div/ul/li[2]/div/ul/li[3]').click()
         time.sleep(2)
         elem = WebDriverWait(driver, 15).until(EC.presence_of_element_located(
             (By.XPATH, "/html/body/form/div[4]/div[2]/div/div/div[2]/div/div/ul/li[3]/div/span")))
         time.sleep(2)
         driver.find_element(By.XPATH,
                             '//*[@id="ContentPlaceHolder1_ASPxPanel2_SearchByAddress1_txtStreetName"]').send_keys(
             STREETNAME)  # passing Street name to driver
         time.sleep(1)
         driver.find_element(By.XPATH,
                             '//*[@id="ContentPlaceHolder1_ASPxPanel2_SearchByAddress1_txtHouseNumber"]').send_keys(
             HOUSENUMBER)  # passing House No  to driver
         time.sleep(1)
         driver.find_element(By.XPATH,
                             '/html/body/form/div[4]/div[2]/div/div/div[3]/div/div[1]/div[2]/div[11]/input').send_keys(
             PIN)  # passing PIN to driver
         time.sleep(1)

         if 'unit' in EXCELADDRESS:  # if in property Address ,unit is present it will enter the Unit address in th driver
             UNIT = str(EXCELADDRESS.split('unit')[-1])
             print(UNIT)

             driver.find_element(By.XPATH,
                                 '/html/body/form/div[4]/div[2]/div/div/div[3]/div/div[1]/div[2]/div[7]/input').send_keys(
                 UNIT)

         driver.find_element(By.XPATH,
                             '/html/body/form/div[4]/div[2]/div/div/div[3]/div/div[1]/div[2]/div[9]/div[1]/input').send_keys(
             City, Keys.ENTER)

         TAXmaincondition = driver.find_element(By.XPATH,
                                                '/html/body/form/div[4]/div[2]/div/div/div[3]/div/div/div[2]/div[1]/span').text

         try:
             if (TAXmaincondition) == 'Found one record matching your search criteria.':  # checking for results

                 time.sleep(4)
                 z = driver.find_element(By.XPATH,
                                         '/html/body/form/div[4]/div[2]/div/div/div[3]/div/div/div[2]/div[2]/table/tbody/tr[2]/td/div/div[2]/div[2]/div[2]/div[2]/span').text  # for checking elements in Tax page
                 # print(text)

                 z1 = (z).split()[0:2]
                 z2 = " ".join(z1)
                 # print(z2.upper())
                 Name = str(dataframe1['NAME'][i]).split()[0:2]
                 Name1 = (" ".join(Name))
                 abc = 10

                 driver.find_element(By.XPATH,
                                     '/html/body/form/div[4]/div[2]/div/div/div[3]/div/div/div[2]/div[2]/table/tbody/tr[2]/td/div/div[2]/div[3]/a').click()
                 text = driver.find_element(By.XPATH,
                                            '/html/body/form/div[4]/div/div/div/div[2]/div[4]/div[1]/div[2]/div/div[2]/span').text  # for getting APN number from Tax page
                 # print(text)

                 os.makedirs(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(int(OrderID)))

                 workbook = openpyxl.load_workbook(os.getcwd() + '\\Input\\' + file)

                 worksheet = workbook.active
                 # print("access to excel sheet")

                 worksheet['C' + str(int(i + 2))] = text
                 worksheet['M1'] = 'GTD'
                 worksheet['N1'] = 'Comments'

                 workbook.save(os.getcwd() + '\\Input\\' + file)
                 # print("saving ")

                 driver.execute_script('window.print();')  # taking print of Tax page
                 time.sleep(3)
                 path = os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(int(OrderID))
                 name = "Tax Sheet"

                 pyautogui.typewrite(path + '\\' + name + '.pdf')
                 time.sleep(2)
                 pyautogui.press('enter')
                 time.sleep(4)
                 driver.close()

                 Code.New_update1_title.Final_A(i, file)

                 # Code.Lien_Report.Final_B(OrderID,F,L,file)

                 Code.BRB_Search.Final_C(OrderID, F, L)

                 workbook1 = openpyxl.Workbook()
                 sheet = workbook1.active
                 # below arrangement of rows for to get search note
                 sheet['A1'] = 'Order Number:'
                 sheet['A2'] = 'BORROWER NAME:'
                 sheet['A3'] = 'ADDRESS:'
                 sheet['A4'] = 'COUNTY:'
                 sheet['A5'] = 'APN:'
                 sheet['A6'] = 'Legal:'
                 sheet['A7'] = 'GTD:'
                 sheet['A8'] = 'NAMES RUN:'
                 sheet['A9'] = '###################################################################################'

                 print("reading excel sheet")
                 df = pd.read_excel(os.getcwd() + '\\Input\\' + file, engine="openpyxl")

                 OrderIDumber = df['Order No'][i]

                 BORROWERNAME = df['NAME'][i]

                 ADDRESS = df['Property Address'][i]

                 COUNTY = df['County Name'][i]

                 APN = df['APN'][i]

                 NAMESRUN1 = df['NAME'][i]

                 GIT = df['GTD'][i]

                 sheet['B1'] = OrderIDumber
                 sheet['B2'] = BORROWERNAME
                 sheet['B3'] = ADDRESS
                 sheet['B4'] = COUNTY
                 sheet['B5'] = APN
                 sheet['B6'] = "(NOT need for IL)"
                 sheet['B7'] = GIT
                 sheet['B8'] = NAMESRUN1
                 sheet['B9'] = "#############"

                 time.sleep(2)
                 workbook1.save(
                     os.getcwd() + '\\Output\\COOK_COUNTY\\' + "Order No " + str(int(OrderID)) + '\\Note.xlsx')

                 df1 = pd.read_excel(
                     os.getcwd() + '\\Output\\COOK_COUNTY\\' + "Order No " + str(int(OrderID)) + '\\filterd_data.xlsx',
                     engine='openpyxl')
                 f = df1[['Doc Number', 'Doc Type', 'Doc Executed', '1st PIN']]

                 df2 = pd.read_excel(
                     os.getcwd() + '\\Output\\COOK_COUNTY\\' + "Order No " + str(int(OrderID)) + '\\Note.xlsx',
                     engine='openpyxl')

                 df_combined = df2._append(f)
                 combinedfile = os.getcwd() + '\\Output\\COOK_COUNTY\\' + "Order No " + str(
                     int(OrderID)) + '\\SearchNoteXL.xlsx'
                 df_combined.to_excel(combinedfile, index=False)


                 end_time = datetime.now()
                 worksheet['k' + str(int(i + 2))] = end_time
                 workbook.save(os.getcwd() + '\\Input\\' + file)

                 # zipping the order output files
                 createZipfile(OrderID)
                 files = [
                     ('UploadFile', (
                     str(OrderID) + ".zip", open(os.getcwd() + '\\Output\\COOK_COUNTY\\' + str(OrderID) + ".zip", 'rb'),
                     'zip'))]
                 # uploading the zipped doc
                 getOrders.uploadDocument(OrderID, OrderNum, "Completed", processId, "Successful", files)
                 source_folder = (os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(OrderID))
                 destination_folder = (os.getcwd() + "\\Processed")

                 shutil.move(source_folder, destination_folder)
                 print("Completed:" + str(int(i) + int(1)))

             else:
                 driver.close()
                 try:
                     os.makedirs(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(OrderID))
                 except Exception:
                     print("Error")
                 workbook = openpyxl.load_workbook(os.getcwd() + '\\Input\\' + file)
                 worksheet = workbook.active
                 worksheet['N1'] = 'Comments'
                 worksheet['N' + str(int(i + 2))] = 'Multiple Property Available'
                 workbook.save(os.getcwd() + '\\Input\\' + file)


         except Exception as e:
             print("Max Retry Error in Tax Page",e)
             getOrders.updateStatus(OrderID, OrderNum, "Exception", processId, e)

             try:
                 os.makedirs(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(OrderID))
             except Exception:
                 print("Error")
             workbook = openpyxl.load_workbook(os.getcwd() + '\\Input\\' + file)
             worksheet = workbook.active
             worksheet['B' + str(int(i + 2))] = 'Max Retry Error in Tax Page/ Recorder Page'
             workbook.save(os.getcwd() + '\\Input\\' + file)
             driver.close()

             # print("Closed")

     except Exception as e:
         print(" Maximum Retry Error.",e)
         getOrders.updateStatus(OrderID, OrderNum, "Exception", processId, e)

# if __name__ == '__main__':
#     file="Cook_county.xlsx"
#     Final_UI(file)

