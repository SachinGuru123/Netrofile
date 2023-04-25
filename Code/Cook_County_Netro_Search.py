def Final_A():

    print("HI")

    import datetime
    import sys
    import threading
    import webbrowser
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    import time
    import pandas as pd
    import shutil, os
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from openpyxl import load_workbook
    from selenium.webdriver.common.action_chains import ActionChains
    from tkinter import Tk
    import tkinter
    import pandas as pd
    import selenium.webdriver
    import requests
    from PIL import Image, ImageTk
    from tkinter import ttk
    import requests
    import shutil
    import threading
    import openpyxl
    from urllib.parse import urljoin
    import requests
    from pyhtml2pdf import converter
    from selenium.webdriver.chrome.options import Options
    import pyautogui
    import glob
    import re,PyPDF2
########################
    folder_path = 'D:\\Title_Files\\Order Sheets'
    pdf_path = glob.glob(os.path.join(folder_path, "*.pdf"))

    workbook = openpyxl.load_workbook('D:\\Title_Files\\Input\\Cook_county.xlsx')
    worksheet = workbook.active

    i = 0
    for path in pdf_path:
        # print(path)

        pdf_file = open(path, 'rb')
        pdf_reader = PyPDF2.PdfReader(pdf_file)

        text = ''
        # for i in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[0]
        text = page.extract_text()

        # print(text)
        with open('D:\\Title_Files\\Order Sheets\\abc.txt', 'w') as f:
            # for line in text:
            f.write(text)

        with open('D:\\Title_Files\\Order Sheets\\abc.txt', 'r') as f:
            text_contents = f.read()

        rows = text_contents.split('\n')
        df = pd.DataFrame(rows)
        df.to_csv('D:\\Title_Files\\Order Sheets\\op.csv', index=False)

        df = pd.read_csv('D:\\Title_Files\\Order Sheets\\op.csv')

        order_numbers = []
        County = []
        Borrower = []
        Address = []
        APN = []

        for index, row in df.iterrows():
            order_number_match = re.search(r'Order Number:\s*(\d{7}-\s?\d+)', row["0"])
            order_number_county = re.search(r'County:\s*\w+\s?\w+', row["0"])
            order_Borrower = re.search(r'((\w+?\s?\w+?\s?\w+\s*)(?=Borrower|BORROWER)|((?<=BORROWER)\s*\w+\s*\w+))',
                                       row["0"])

            order_Address = re.search(r'Address:\s*(.*)', row["0"])

            if re.search(r'PIN/APN:\s*(.*)', row["0"]):
                order_APN = re.search(r'PIN/APN:\s*(.*)', row["0"])  # PIN/APN :
                APN.append(order_APN.group(1))
                APN_NUMBER = (APN[0].split(' ')[2])  # APN_NUMBER = (APN[0].split(' ')[2])
                print(APN_NUMBER)
                worksheet['B' + str(int(i + 2))] = APN_NUMBER

            if re.search(r'PIN/APN :\s*(.*)', row["0"]):
                order_APN = re.search(r'PIN/APN :\s*(.*)', row["0"])  # PIN/APN :
                APN.append(order_APN.group(1))
                APN_NUMBER = (APN[0].split(' ')[1])  # APN_NUMBER = (APN[0].split(' ')[2])
                print(APN_NUMBER)
                worksheet['B' + str(int(i + 2))] = APN_NUMBER

            if order_number_match:
                order_numbers.append(order_number_match.group(1))

            if order_number_county:
                County.append(order_number_county.group(0))

            if order_Borrower:
                Borrower.append(order_Borrower.group(0))

            if order_Address:
                Address.append(order_Address.group(0))

        ORDER_NUMBER = (order_numbers[0])
        COUNTY = (County[0].strip('County:'))  # [2])#.strip('County:'))

        print(COUNTY)

        print(ORDER_NUMBER)

        BORROWER_NAME = (Borrower[0])  # .strip("BORROWER"))
        print(BORROWER_NAME)

        ADDRESS = (Address[0].strip('Address:'))

        worksheet['A' + str(int(i + 2))] = ORDER_NUMBER
        worksheet['E' + str(int(i + 2))] = COUNTY
        worksheet['C' + str(int(i + 2))] = BORROWER_NAME
        worksheet['D' + str(int(i + 2))] = ADDRESS
        # if re.search(r'PIN/APN:\s*(.*)', row["0"]):
        # worksheet['E' + str(int(i + 2))] = APN_NUMBER
        workbook.save('D:\\Title_Files\\Input\\Cook_county.xlsx')
        i = i + 1
#############################
    dataframe1 = pd.read_excel(
        'D:\\Title_Files\\Input\\Cook_county.xlsx')

    E = dataframe1[dataframe1.columns[0]].count()
    for i in range(E):
     if dataframe1['County Name'][i].strip()==('Cook'):

            EXCELAPN = str(dataframe1['APN'][i].replace("-",''))
            print(EXCELAPN)

            EXCELNAME = dataframe1['NAME'][i]

            EXCELORDERNO = dataframe1['Order No'][i].replace("-",'')

            #order number is created
            os.makedirs("D:\\Title_Files\\Output\\COOK_COUNTY\\" + "Order No " + str(int(EXCELORDERNO)))

            chrome_driver = 'D:\chromedriver_win32\chromedriver.exe'

            chrome_options = Options()
            chrome_options.add_argument('--kiosk-printing')
            chrome_options.add_argument('--disable-extensions')

            driver = webdriver.Chrome(options=chrome_options)



            driver.maximize_window()
            driver.get('https://crs.cookcountyclerkil.gov/Search/Additional')
            #driver.maximize_window()
            time.sleep(2)
            driver.find_element(By.XPATH, '/html/body/div[2]/div/div[3]/div/div/div[3]/div[2]/h2/button').click()
            time.sleep(3)
            # driver.find_element(By.XPATH, '//*[@id="RecordedFromDate"]').send_keys(StartDate)
            # driver.find_element(By.XPATH, '//*[@id="RecordedToDate"]').send_keys(EndDate)

            driver.find_element(By.XPATH,
                                '/html/body/div[2]/div/div[3]/div/div/div[3]/div[2]/div/div/form/div[1]/div/input').send_keys(
                EXCELNAME, Keys.ENTER)

            #converter.convert(driver., "D:\Title_Files\PN Results.pdf")
            a = driver.find_element(By.XPATH, '//table')
            df = pd.read_html(a.get_attribute('outerHTML'))[0]

            #to take print of Index Value
            driver.execute_script('window.print();')
            time.sleep(4)
            path="D:\\Title_Files\\Output\\COOK_COUNTY\\" + "Order No "+str(EXCELORDERNO)
            name="Index"
            pyautogui.typewrite(path)
            time.sleep(1)
            pyautogui.press('enter')
            pyautogui.press('enter')
            pyautogui.press('enter')
            pyautogui.press('enter')
            pyautogui.press('enter')
            pyautogui.press('enter')
            time.sleep(3)
            pyautogui.press('enter')
            pyautogui.press('enter')
            pyautogui.typewrite(name)
            pyautogui.press('enter')
            time.sleep(2)
            print("Done")
            #time.sleep()

            # need to un comment
            #q = driver.find_element(By.XPATH,'/html/body/div[2]/div/div[3]/div/form/div[5]/div[4]/table/tbody/tr[2]/td[11]/a').get_attribute('href')
            #height = driver.execute_script("return document.body.scrollHeight")
            #width = driver.execute_script("return document.body.scrollWidth")

            #driver.set_window_size(width, height)
            #screenshot = driver.find_element(By.TAG_NAME, 'body').screenshot_as_png

            #with open("D:\\Title_Files\\image.png", 'wb') as f:
                #f.write(screenshot)

            a = driver.find_element(By.XPATH, '//table')
            df = pd.read_html(a.get_attribute('outerHTML'))[0]

            # print(q.split("=",4))

            driver.get('https://crs.cookcountyclerkil.gov/Search/ResultByPin?id1=' +str(EXCELAPN))
            print("Added")

            time.sleep(2)
            COunt = driver.find_element(By.XPATH, '//*[@id="result"]/div[1]/div/span').text
            print(str(COunt))
            COunt = int(COunt) + 1
            print(COunt)


            # start_time = datetime.datetime.now()
            # print(start_time)
            workbook = openpyxl.load_workbook(
                'D:\\Title_Files\\Input\\Cook_county.xlsx')
            worksheet = workbook.active
            start_time = datetime.datetime.now()
            worksheet['F' + str(int(i + 2))] = start_time
            k = 1
            #os.makedirs("D:\\Title_Files\\Output\\COOK_COUNTY\\" +"Order No " + str(int(EXCELORDERNO)))
            #converter.convert(p,"D:\\Title_Files\\Output\\COOK_COUNTY\\" + "Order No " + str(int(EXCELORDERNO)) + '\\Index Results.pdf')
            df.to_excel("D:\\Title_Files\\Output\\COOK_COUNTY\\" + "Order No "+str(int(EXCELORDERNO))+'\\Index Results.xlsx', index=False)
            #converter.convert(driver.current_url, "D:\Title_Files\Output\COOK_COUNTY\Order No"+ str(int(EXCELORDERNO))+"\APN Results.pdf")

            while k < int(COunt):
                try:
                    # if k < int(COunt):

                    if bool(driver.find_element(By.XPATH,
                                                '//*[@id="result"]/div[4]/table/tbody/tr[' + str(
                                                    k) + ']/td[2]/a')) == True:

                        Q = driver.find_element(By.XPATH,
                                                '/html/body/div[2]/div/div[3]/div/form/div[3]/div[4]/table/tbody/tr[' + str(
                                                    k) + ']/td[3]/span').text
                        print(Q)
                        # worksheet['G' + str(int(k + 1))]=Q
                        d = driver.find_element(By.XPATH,
                                                '/html/body/div[2]/div/div[3]/div/form/div[3]/div[4]/table/tbody/tr[' + str(
                                                    k) + ']/td[2]/a').get_attribute('href')
                        print(d)
                        driver.get(d)

                        A = driver.find_element(By.XPATH,
                                                '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a').get_attribute(
                            'href')
                        pdf_url = A
                        #os.makedirs("D:\\Title_Files\\Output\\COOK_COUNTY\\"+str(EXCELAPN))
                        try:
                            url = A
                            r = requests.get(url, stream=True)
                            print(r.status_code)
                            with open(
                                    'D:\\Title_Files\\Output\\COOK_COUNTY\\' + "Order No "+str(int(EXCELORDERNO))
                                         + '\\Doc' + str(k) + '   ' + str(Q) + '.pdf', 'wb') as fd:
                                for chunk in r.iter_content(chunk_size=20):
                                    fd.write(chunk)

                        finally:
                            url = A
                            r = requests.get(url, stream=True)
                            print(r.status_code)
                            with open(
                                    'D:\\Title_Files\\Output\\COOK_COUNTY\\' + "Order No " +str(int(EXCELORDERNO))
                                         + '\\Doc' + str(k) + '   ' + str(Q) + '.pdf', 'wb') as fd:
                                for chunk in r.iter_content(chunk_size=20):
                                    fd.write(chunk)

                        k = k + 1
                        print(k)
                        driver.back()

                except Exception as Err:
                    print(Err)
                    k = k + 1
                    driver.back()

            end_time = datetime.datetime.now()
            worksheet['G' + str(int(i + 2))] = end_time
            worksheet['H' + str(int(i + 2))] = "Completed"
            #worksheet['F' + str(int(i + 2))] = str(int(COunt)-int(1)) #-int(1)
            worksheet['I' + str(int(i + 2))]=str(end_time-start_time)

            workbook.save('D:\\Title_Files\\Input\\Cook_county.xlsx')
    source = 'D:\\Title_Files\\Input\\Cook_county.xlsx'
    destination = 'D:\\Title_Files\\Processed\\Cook_county.xlsx'
    shutil.move(source, destination)

    print("Completed")


    driver.close()



if __name__=='__main__':
    Final_A()
