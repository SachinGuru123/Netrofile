from selenium.common import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time

import shutil, os

import pandas as pd

import openpyxl
from urllib.parse import urljoin
import requests
from pyhtml2pdf import converter
from selenium.webdriver.chrome.options import Options
import pyautogui
import glob
import re, PyPDF2
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import math

def Final_A(i,file):

    time.sleep(1)
    par_dir = os.getcwd()
    dataframe1 = pd.read_excel(par_dir+'\\Input\\'+file)
    

    E = dataframe1[dataframe1.columns[0]].count()
    #print(dataframe1['County Name'][0])
    print("Running Recorder Page")


    if dataframe1['County Name'][i] == ('Cook'):
    
        EXCELAPN = str(dataframe1['APN'][i])  # getting APN number
    
        EXCELNAME = dataframe1['NAME'][i]  # getting Name
        splitted_name = EXCELNAME.split()[-1]
    
        ExCELADDRESS = dataframe1['Property Address'][i]  # getting Address
        aa = ExCELADDRESS.split()[0:3]
        ab = (' '.join(map(str, aa)))
        # print(splitted_name)
        # print(ab)
        ############################################################################
        EXCELORDERNO = dataframe1['Order ID'][i]  # getting Order No
    
        chrome_driver = 'chromedriver_win32\chromedriver.exe'
        time.sleep(1)
        chrome_options = Options()
        chrome_options.add_argument('--kiosk-printing')
        chrome_options.add_argument('--disable-extensions')
        time.sleep(1)
        driver = webdriver.Chrome(options=chrome_options)
        time.sleep(3)
        driver.get('https://crs.cookcountyclerkil.gov/Search')  # opening Tax page
        driver.maximize_window()
        time.sleep(1)
    
        driver.find_element(By.XPATH,
                            '/html/body/div[2]/div/div[3]/div/div/form/div[2]/div[2]/div[3]/div/div[2]/input').send_keys(
            splitted_name + " " + ab, Keys.ENTER)  # passing Borrower name to chrome driver
    
        aq = driver.find_element(By.XPATH,
                                 '/html/body/div[2]/div/div[3]/div/form[1]/div[2]/div/div[2]').text  # for extracting GTD from chrome driver
    
        workbook = openpyxl.load_workbook(os.getcwd() + '\\Input\\' + file)
        worksheet = workbook.active
        GIT = aq.split("|")[0]
        worksheet['M' + str(int(i + 2))] = GIT  # passing GIT to Excel
        workbook.save(os.getcwd() + '\\Input\\' + file)
    
        a = driver.find_element(By.XPATH, '//table')  # Extracting Table in recorder page
        df = pd.read_html(a.get_attribute('outerHTML'))[0]
    
        # to take print of Index Value of Recorder Page
        driver.execute_script('window.print();')
        time.sleep(5)
    
        path = os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO)
        time.sleep(3)
        name = "Name Index"
        pyautogui.FAILSAFE = False
        pyautogui.typewrite(path + '\\' + name + '.pdf')
        pyautogui.press('enter')
        time.sleep(1)
    
        a = driver.find_element(By.XPATH, '//table')  # extracting Tabular Data
        df = pd.read_html(a.get_attribute('outerHTML'))[0]
    
        time.sleep(2)
        COunt = driver.find_element(By.XPATH, '//*[@id="result"]/div[1]/div/span').text
    
        COunt = int(COunt) + 1
    
        workbook = openpyxl.load_workbook(
            os.getcwd() + '\\Input\\' + file)
        worksheet = workbook.active
    
        k = 1  # for DOC counting Reference
    
        df.to_excel(
            os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO) + '\\Name Results.xlsx',
            index=False)
        workbook.save(os.getcwd() + '\\Input\\' + file)
    
        j = 1
        link = []  # passsing all href links of recorder pages to list
        while j < int(COunt):
            href_ad = driver.find_element(By.XPATH,
                                          '//*[@id="tblData"]/tbody/tr[' + str(j) + ']/td[2]/a').get_attribute(
                "href")
            # print(href_ad)
            link.append(href_ad)
            j += 1
    
        N = 1
        for k in link:  # opening all href by passing each value to driver
    
            driver.get(k)
            WebDriverWait(driver, 15).until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a')))
            d = driver.find_element(By.XPATH,
                                    '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[1]/div[1]/table/tbody/tr[1]/td').text
            LinkF = driver.find_element(By.XPATH,
                                        '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a').get_attribute(
                'href')
    
            max_retry = 5
            retry_count = 0
            while retry_count < max_retry:  # to download the pdf Document file
                try:
    
                    r = requests.get(LinkF)
                    time.sleep(1)
    
                    with open('Output\\COOK_COUNTY\\' + "Order No " + str(EXCELORDERNO)
                              + '\\Doc' + str(N) + '   ' + str(d) + '.pdf', 'wb') as fd:
                        for chunk in r.iter_content(chunk_size=40):
                            fd.write(chunk)
    
                    break
    
                except Exception as e:
                    retry_count += 1
    
            N += 1
    
        driver.get('https://crs.cookcountyclerkil.gov/Search')  # opening for APN Search recorder Page
    
        driver.find_element(By.XPATH,
                            '/html/body/div[2]/div/div[3]/div/div/form/div[2]/div[2]/div[3]/div/div[2]/input').clear()  # to clear the input Value box in Recorder Page
        driver.find_element(By.XPATH,
                            '/html/body/div[2]/div/div[3]/div/div/form/div[2]/div[2]/div[3]/div/div[2]/input').send_keys(
            EXCELAPN, Keys.ENTER)  # passing APN Number to Driver
        time.sleep(3)
        aa = driver.find_element(By.XPATH, '//table')
        df1 = pd.read_html(aa.get_attribute('outerHTML'))[0]
        df1.to_excel(
            os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO) + '\\APN Results.xlsx',
            index=False)  # to save APN page Tabular Data
    
        ########################################################################
        workbook = openpyxl.load_workbook(
            os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO) + '\\APN Results.xlsx')
        worksheet = workbook.active
    
    DocCOunt = driver.find_element(By.XPATH, '//*[@id="result"]/div[1]/div/span').text
    
    j = 1
    
    while j <= int(DocCOunt):
        href_ad = driver.find_element(By.XPATH,
                                      '//*[@id="tblData"]/tbody/tr[' + str(j) + ']/td[2]/a').get_attribute(
            "href")
    
        worksheet['K' + str(int(j + 1))] = href_ad
        j += 1
    
    workbook.save(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO) + '\\APN Results.xlsx')
    
    #########################################################################
    df2 = pd.read_excel(
        os.getcwd() + '\\Output\\COOK_COUNTY\\' + "Order No " + str(EXCELORDERNO) + '\\Name Results.xlsx')
    lastdate = (df2['Doc Recorded'].iloc[-1])
    
    data_frame = pd.read_excel(
        os.getcwd() + '\\Output\\COOK_COUNTY\\' + "Order No " + str(EXCELORDERNO) + '\\APN Results.xlsx')
    
    data_frame['Doc Recorded'] = pd.to_datetime(data_frame['Doc Recorded'], format='%m/%d/%Y')
    
    filterd_data = data_frame[data_frame['Doc Recorded'] >= lastdate]
    
    filterd_data.to_excel(
        os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO) + "\\filterd_data.xlsx",
        index=False)
    
    df2 = pd.read_excel(
        os.getcwd() + '\\Output\\COOK_COUNTY\\' + "Order No " + str(EXCELORDERNO) + '\\Name Results.xlsx')
    df3 = pd.read_excel(
        os.getcwd() + '\\Output\\COOK_COUNTY\\' + "Order No " + str(EXCELORDERNO) + '\\filterd_data.xlsx')
    
    col1 = df2['Doc Number']
    col2 = df3['Doc Number']
    
    list = []
    # to remove the Duplicate Numbers in Name result and filtered_data Excel
    mask = ~col1.isin(col2)
    non_matching_values = col1[~mask]
    non_comapred_values = col2[~col2.isin(col1)]  # Doc Number in  which needs to be download in APN page
    
    lastdate = (df2['Doc Recorded'].iloc[-1])
    
    L = []  # creating list for passing condition satisfied document number
    for x in non_comapred_values:
        L.append(x)  # passing condition satisfied document number to list for to download
    
    Total_files_No = driver.find_element(By.XPATH,
                                         '/html/body/div[2]/div/div[3]/div/form[2]/div[4]/div[1]/div/span').text  # No of Docs Available in recorder Page
    driver.execute_script('window.print();')
    time.sleep(4)
    # print(Total_files_No)
    path = os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO)
    name = "ParcelNumber_index"
    time.sleep(1)
    pyautogui.typewrite(path + '\\' + name + '.pdf')
    time.sleep(1)
    pyautogui.press('enter')
    time.sleep(1)
    
    column_to_compare = 'Doc Number'  # compairing the NameDoc and FileterDoc to download files in Dataframe
    
    combined_df = pd.concat([df2, df3])
    duplicated_df = combined_df.drop_duplicates(subset=column_to_compare)
    duplicated_df.to_excel(
        os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO) + "\\filterd_data1.xlsx",
        index=False)
    
    url_list = duplicated_df['Unnamed: 10'].to_list()
    url_list = [url for url in url_list if isinstance(url, str) and url.strip()]
    
    for url in url_list:  # for downloading the Docs in APN page which met the given condtions
        driver.get(url)
        Doc_num_APN_PAGE = driver.find_element(By.XPATH, '//*[@id="divcol1"]/div[1]/table/tbody/tr[1]/td').text
    
        url = driver.find_element(By.XPATH,
                                  '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a').get_attribute(
            'href')  # to get final href link to downlad pdf document
    
        max_retry = 5
        retry_count = 0
        # Documents downloading part
        while retry_count < max_retry:
            try:
                r = requests.get(url)
                time.sleep(2)  # verify=False
                with open(
                        os.getcwd() + '\\Output\\COOK_COUNTY\\' + "Order No " + str(
                            EXCELORDERNO)
                        + '\\APN Doc' + " " + str(int(Doc_num_APN_PAGE)) + " " + '.pdf', 'wb') as f:
                    for chunk in r.iter_content(chunk_size=40):
                        f.write(chunk)
                    retry_count += 1
    
                    break
    
            except Exception as e:
                retry_count += 1
    
    print("APN SITE COMPLETED")
    
    #############################################Second Borrower name Search ###################
    
    workbook = openpyxl.load_workbook('Input\\' + file)
    sheet = workbook.active
    
    cell_value = sheet['I' + str(i + 2)].value
    # print(cell_value)
    
    if cell_value is None:
        print("Second name not Exist")
    
    
    else:
        # if second name available in Excel its comes to else block.
        driver.get('https://crs.cookcountyclerkil.gov/Search')
        time.sleep(5)
    
        # driver.find_element(By.XPATH,
        # '/html/body/div[2]/div/div[3]/div/div/form/div[2]/div[2]/div[3]/div/div[2]/input').clear()
        driver.find_element(By.XPATH,
                            '/html/body/div[2]/div/div[3]/div/div/form/div[2]/div[2]/div[3]/div/div[2]/input').send_keys(
            cell_value, Keys.ENTER)
    
        z = driver.find_element(By.XPATH, '//*[@id="result"]/div[1]/div/span[1]').text
        z = (z.replace(",", ""))
        total = int(z)  # total no of docs in recorder page
        df1 = pd.DataFrame()
        # print(total)
    
        Cnumber = driver.find_element(By.XPATH,
                                      '/html/body/div[2]/div/div[3]/div/form[2]/div[4]/div[1]/div/span[1]').text  # total no of docs in the Name search Page
    
        if total >= int(100):
            qq = driver.find_element(By.XPATH, '//*[@id="Paging"]/div/ul/li[2]/a').get_attribute("href")
            qq = qq[:len(qq) - 1]
            # print(qq)
            time.sleep(2)
            a = driver.find_element(By.XPATH, '//table')
            df = pd.read_html(a.get_attribute('outerHTML'))[0]
            df1 = df1.append(df)
            time.sleep(2)
            r = requests.get(qq)
            time.sleep(2)
            soup = BeautifulSoup(r.text, 'lxml')
    
            df1 = pd.DataFrame()
            j = 1
    
            CC = int(total) / 100
            roundoff = round(CC)
            # print(roundoff)
    
            while int(j) <= int(roundoff):
                Np = soup.find("a", rel='next').get("href")
                cnp = "https://crs.cookcountyclerkil.gov" + Np
                url = cnp
                # r = requests.get(url,verify=False)
                time.sleep(1)
                # updated = url[:-1]
                updated = url.rstrip('0123456789')  # to remove page numbers from the href link
                # print(j)
                updated = updated + str(j)
    
                max_retry = 2
                retry_count = 0
                while retry_count < max_retry:
                    try:
                        r = requests.get(updated)  # verify=False
                        # print(r.content)
                        # print(updated)
                        break
                        # retry_count += 1
    
                    except Exception as e:
                        retry_count += 1
    
                soup = BeautifulSoup(r.content, 'html.parser')
                table = soup.find('table')
                df = pd.read_html(str(table))[0]
                # print(df)
                df1 = df1.append(df)
                # print(df1)
    
                j += 1
    
            df1.to_excel(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                EXCELORDERNO) + '\\DataExtraction1.xlsx', index=False)
    
            df = pd.read_excel(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                EXCELORDERNO) + '\\DataExtraction1.xlsx')
    
            df['APN Number'] = df['1st PIN'].str.split(' ').str[0]
            APN = str(EXCELAPN)  # '02-08-400-010-0000'
            indices = df.loc[df['APN Number'] == APN].index.tolist()
            a = []
            # print indices of all matches one by one
            for i in indices:
                if i / 100 < 1:
                    # print("Index: 1")
                    a.append(int(1))
                else:
                    # print("Index:", round(i / 100))
                    a.append(round(i / 100))
            # In below condition used to find out APN number available page
            unique_numbers = []
            [unique_numbers.append(num) for num in a if num not in unique_numbers]
            # print(unique_numbers)
    
            l = 1
    
            for k in unique_numbers:
                driver.get(qq + str(k))
                driver.execute_script('window.print();')
                time.sleep(5)
                pyautogui.press('enter')
                path = os.getcwd() + '\\Output\\COOK_COUNTY\\' + "Order No " + str(
                    EXCELORDERNO) + '\\SecondName Search' + str(l) + '.pdf'
                pyautogui.typewrite(path)
    
                pyautogui.press('enter')
                time.sleep(8)
                l += 1
    
            #####################Lien Part Added##############################Start
            driver.close()
            print("line 436")
    
            pattern = r'\d{2}-\d{2}-\d{4}'
            match = re.search(pattern, aq)
            Splitting = (match.group())
            Replaced = Splitting.replace('-', '/')
            print(Replaced)
    
            print("line 438")
            Effective_Date = datetime.strptime(Replaced, '%m/%d/%Y')
            print(Effective_Date)
            print("line 440")
    
            STATE_LEIN_DATE = Effective_Date - timedelta(days=365 * 20)  # Qualifying Date For Sate Tax
            FED_TAX_DATE = Effective_Date - timedelta(days=365 * 10)  # Qualifying Date For Fed Tax
            UCC_DATE = Effective_Date - timedelta(days=365 * 5)  # Qualifying Date For UCC
            JDG_DATE = Effective_Date - timedelta(days=365 * 20)  # Qualifying Date For Judgemnt
            HOA_DATE = Effective_Date - timedelta(days=365 * 5)  ##Qualifying Date For HOA
            print("line 446")
            print(df)
            df['Doc Recorded'] = pd.to_datetime(df['Doc Recorded'], format='%m/%d/%Y')
            print("line 458")
            Liens = df['Doc Type'].str.contains('JUDGMENT|LIEN|STATE LIEN|FEDERAL LIEN|UCC|HOA',
                                                case=False, regex=True)
            filterd_df = df[Liens]
            print("line 453")
            FED = filterd_df[
                ((filterd_df['Doc Type'] == 'FEDERAL LIEN') & (filterd_df['Doc Recorded'] >= FED_TAX_DATE))]
            ST = filterd_df[
                ((filterd_df['Doc Type'] == 'STATE LIEN') & (filterd_df['Doc Recorded'] >= STATE_LEIN_DATE))]
            UCC = filterd_df[((filterd_df['Doc Type'] == 'UCC') & (filterd_df['Doc Recorded'] >= UCC_DATE))]
            JUD = filterd_df[
                ((filterd_df['Doc Type'] == 'JUDGMENT') & (filterd_df['Doc Recorded'] >= JDG_DATE))]
            HOA = filterd_df[((filterd_df['Doc Type'] == 'HOA') & (filterd_df['Doc Recorded'] >= HOA_DATE))]
            print("line 471")
            with pd.ExcelWriter(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                    EXCELORDERNO) + '\\LeinDocSN.xlsx') as writer:
                FED.to_excel(writer, sheet_name='Sheet1', index=False, startrow=0)
                ST.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(FED) + 1, header=False)
                UCC.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(ST) + 1, header=False)
                JUD.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(UCC) + 1, header=False)
                HOA.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(JUD) + 1, header=False)
    
            print("line 480")
            dataframe2 = pd.read_excel(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                EXCELORDERNO) + '\\LeinDocSN.xlsx')
    
            chrome_driver = 'chromedriver_win32\chromedriver.exe'
            time.sleep(1)
            chrome_options = Options()
            chrome_options.add_argument('--kiosk-printing')
            chrome_options.add_argument('--disable-extensions')
            time.sleep(2)
            driver = webdriver.Chrome(options=chrome_options)
            time.sleep(2)
            print("line 492")
        for Document_Number in (dataframe2['Doc Number']):
            time.sleep(1)
            driver.get('https://crs.cookcountyclerkil.gov/Search')
            driver.find_element(By.XPATH,
                                '/html/body/div[2]/div/div[3]/div/div/form/div[2]/div[2]/div[3]/div/div[2]/input').send_keys(
                Document_Number, Keys.ENTER)
            driver.maximize_window()
            time.sleep(5)
            h = driver.find_element(By.XPATH, '//*[@id="tblData"]/tbody/tr/td[2]/a').get_attribute('href')
            driver.get(h)
    
            WebDriverWait(driver, 15).until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a')))
            LienType = driver.find_element(By.XPATH, '//*[@id="divcol1"]/div[1]/table/tbody/tr[2]/td').text
            LienNumber = driver.find_element(By.XPATH,
                                             '//*[@id="divcol1"]/div[1]/table/tbody/tr[1]/td').text
            LinkF = driver.find_element(By.XPATH,
                                        '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a').get_attribute(
                'href')
    
            max_retry = 5
            retry_count = 0
            while retry_count < max_retry:  # to download the pdf Document file
                try:
                    time.sleep(2)
                    r = requests.get(LinkF)
                    time.sleep(1)
    
                    with open(
                            os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                                EXCELORDERNO) + '\\' + LienType + ' ' + LienNumber + '.pdf',
                            'wb') as fd:
                        for chunk in r.iter_content(chunk_size=40):
                            fd.write(chunk)
    
                    break
    
                except Exception as e:
                    retry_count += 1
        #####################Lien Part Added##############################End
        else:
            secNameExtraction = driver.find_element(By.XPATH, '//table')
            df1 = pd.read_html(secNameExtraction.get_attribute('outerHTML'))[0]
            df1.to_excel(
                os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                    EXCELORDERNO) + '\\DataExtraction1.xlsx',
                index=False)  # to save Second Name Tabular Data
    
            # Below condition is used to take print of page in which Doc number is less than 100
            driver.execute_script('window.print();')
            time.sleep(5)
            path = os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO)
            name = "Full_Name_search_Index2"
            # pyautogui.FAILSAFE = False
            pyautogui.typewrite(path + '\\' + name + '.pdf')
            time.sleep(2)
            pyautogui.press('enter')
            time.sleep(4)
    
            #####################Lien Part Added##############################Start
            driver.close()
            print("line 553")
    
            pattern = r'\d{2}-\d{2}-\d{4}'
            match = re.search(pattern, aq)
            Splitting = (match.group())
            Replaced = Splitting.replace('-', '/')
            print(Replaced)
    
            print("line 561")
            Effective_Date = datetime.strptime(Replaced, '%m/%d/%Y')
            print(Effective_Date)
            print("line 564")
    
            STATE_LEIN_DATE = Effective_Date - timedelta(days=365 * 20)  # Qualifying Date For Sate Tax
            FED_TAX_DATE = Effective_Date - timedelta(days=365 * 10)  # Qualifying Date For Fed Tax
            UCC_DATE = Effective_Date - timedelta(days=365 * 5)  # Qualifying Date For UCC
            JDG_DATE = Effective_Date - timedelta(days=365 * 20)  # Qualifying Date For Judgemnt
            HOA_DATE = Effective_Date - timedelta(days=365 * 5)  ##Qualifying Date For HOA
            print("line 571")
    
            df1['Doc Recorded'] = pd.to_datetime(df1['Doc Recorded'], format='%m/%d/%Y')
            print("line 574")
            Liens = df1['Doc Type'].str.contains('JUDGMENT|LIEN|STATE LIEN|FEDERAL LIEN|UCC|HOA',
                                                 case=False, regex=True)
            filterd_df = df1[Liens]
            print("line 587")
            FED = filterd_df[
                ((filterd_df['Doc Type'] == 'FEDERAL LIEN') & (filterd_df['Doc Recorded'] >= FED_TAX_DATE))]
            ST = filterd_df[
                ((filterd_df['Doc Type'] == 'STATE LIEN') & (filterd_df['Doc Recorded'] >= STATE_LEIN_DATE))]
            UCC = filterd_df[((filterd_df['Doc Type'] == 'UCC') & (filterd_df['Doc Recorded'] >= UCC_DATE))]
            JUD = filterd_df[
                ((filterd_df['Doc Type'] == 'JUDGMENT') & (filterd_df['Doc Recorded'] >= JDG_DATE))]
            HOA = filterd_df[((filterd_df['Doc Type'] == 'HOA') & (filterd_df['Doc Recorded'] >= HOA_DATE))]
            print("line 587")
            with pd.ExcelWriter(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                    EXCELORDERNO) + '\\LeinDocSN.xlsx') as writer:
                FED.to_excel(writer, sheet_name='Sheet1', index=False, startrow=0)
                ST.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(FED) + 1, header=False)
                UCC.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(ST) + 1, header=False)
                JUD.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(UCC) + 1, header=False)
                HOA.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(JUD) + 1, header=False)
    
            print("line 596")
            dataframe2 = pd.read_excel(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                EXCELORDERNO) + '\\LeinDocSN.xlsx')
    
            chrome_driver = 'chromedriver_win32\chromedriver.exe'
            time.sleep(1)
            chrome_options = Options()
            chrome_options.add_argument('--kiosk-printing')
            chrome_options.add_argument('--disable-extensions')
            time.sleep(2)
            driver = webdriver.Chrome(options=chrome_options)
            time.sleep(2)
            print("line 492")
            for Document_Number in (dataframe2['Doc Number']):
                time.sleep(1)
                driver.get('https://crs.cookcountyclerkil.gov/Search')
                driver.find_element(By.XPATH,
                                    '/html/body/div[2]/div/div[3]/div/div/form/div[2]/div[2]/div[3]/div/div[2]/input').send_keys(
                    Document_Number, Keys.ENTER)
                driver.maximize_window()
                time.sleep(5)
                h = driver.find_element(By.XPATH, '//*[@id="tblData"]/tbody/tr/td[2]/a').get_attribute('href')
                driver.get(h)
    
                WebDriverWait(driver, 15).until(EC.presence_of_element_located(
                    (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a')))
                LienType = driver.find_element(By.XPATH, '//*[@id="divcol1"]/div[1]/table/tbody/tr[2]/td').text
                LienNumber = driver.find_element(By.XPATH,
                                                 '//*[@id="divcol1"]/div[1]/table/tbody/tr[1]/td').text
                LinkF = driver.find_element(By.XPATH,
                                            '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a').get_attribute(
                    'href')
    
                max_retry = 5
                retry_count = 0
                while retry_count < max_retry:  # to download the pdf Document file
                    try:
                        time.sleep(2)
                        r = requests.get(LinkF)
                        time.sleep(1)
    
                        with open(
                                os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                                    EXCELORDERNO) + '\\' + LienType + ' ' + LienNumber + '.pdf',
                                'wb') as fd:
                            for chunk in r.iter_content(chunk_size=40):
                                fd.write(chunk)
    
                        break
    
                    except Exception as e:
                        retry_count += 1
            #####################Lien Part Added##############################End
    
    ###################################################first borrower name part#######################
    
    time.sleep(4)
    driver.get('https://crs.cookcountyclerkil.gov/Search')
    time.sleep(4)
    
    # driver.find_element(By.XPATH,
    # '/html/body/div[2]/div/div[3]/div/div/form/div[2]/div[2]/div[3]/div/div[2]/input').clear()
    driver.find_element(By.XPATH,
                        '/html/body/div[2]/div/div[3]/div/div/form/div[2]/div[2]/div[3]/div/div[2]/input').send_keys(
        EXCELNAME, Keys.ENTER)
    
    z = driver.find_element(By.XPATH, '//*[@id="result"]/div[1]/div/span[1]').text
    z = (z.replace(",", ""))
    total = int(z)  # (int(z) // int(100) - 1)
    df1 = pd.DataFrame()
    # print(total)
    
    Cnumber1 = driver.find_element(By.XPATH,
                                   '/html/body/div[2]/div/div[3]/div/form[2]/div[4]/div[1]/div/span[1]').text
    # print(Cnumber1)
    
    if total >= int(100):
    
        qq = driver.find_element(By.XPATH, '//*[@id="Paging"]/div/ul/li[2]/a').get_attribute("href")
        qq = qq[:len(qq) - 1]
        # print(qq)
        time.sleep(2)
        a = driver.find_element(By.XPATH, '//table')
        df = pd.read_html(a.get_attribute('outerHTML'))[0]
        # print(df)
        df1 = df1.append(df)
        # df1= df1.append(df)
    
        r = requests.get(qq)
        time.sleep(2)
        soup = BeautifulSoup(r.text, 'lxml')
    
        df1 = pd.DataFrame()
        j = 1
    
        CC = int(total) / 100
        roundoff1 = round(CC)
        # print(roundoff1)
    
        while int(j) < int(roundoff1):
            Np = soup.find("a", rel='next').get("href")
            cnp = "https://crs.cookcountyclerkil.gov" + Np
            url = cnp
            # r = requests.get(url,verify=False)
            time.sleep(1)
            # updated = url[:-1]
            updated = url.rstrip('0123456789')
            # print(j)
            updated = updated + str(j)
            # updated = url.rstrip('0123456789')
            # print(updated)
    
            max_retry = 3
            retry_count = 0
            while retry_count < max_retry:
                try:
                    r = requests.get(updated)  # verify=False
                    # print(r.content)
                    # print(updated)
                    retry_count += 1
                    break
    
                except Exception as e:
                    retry_count += 1
    
            soup = BeautifulSoup(r.content, 'html.parser')
            table = soup.find('table')
            time.sleep(3)
            df = pd.read_html(str(table))[0]
            time.sleep(1)
            # print(df)
            df1 = df1.append(df)
            # print(df1)
    
            j += 1
            # print(j)
    
        # df1.to_excel('D:\Title_Files\Output\COOK_COUNTY\Order No 1192237\c.xlsx', index=False)  # index=False
        df1.to_excel(
            os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO) + '\\DataExtraction.xlsx',
            index=False)
    
        df = pd.read_excel(
            os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO) + '\\DataExtraction.xlsx',
            engine='openpyxl')
    
        df['APN Number'] = df['1st PIN'].str.split(' ').str[0]
        APN = str(EXCELAPN)  # '02-08-400-010-0000'
        indices = df.loc[df['APN Number'] == APN].index.tolist()
    
        a = []
        # print indices of all matches one by one
        for w in indices:
            if w / 100 < 1:
                # print("Index: 1")
                a.append(int(1))
            else:
                # print("Index:", round(w / 100))
                a.append(round(w / 100))
    
        unique_numbers = []
        [unique_numbers.append(num) for num in a if num not in unique_numbers]
        # print(unique_numbers)
    
        l = 1
        for k in unique_numbers:
            driver.get(qq + str(k))
            driver.execute_script('window.print();')
            time.sleep(5)
            pyautogui.press('enter')
            path = os.getcwd() + '\\Output\\COOK_COUNTY\\' + "Order No " + str(
                EXCELORDERNO) + '\\Name Search' + str(l) + '.pdf'
            pyautogui.typewrite(path)
    
            pyautogui.press('enter')
            time.sleep(8)
            l += 1
    
            #####################Lien Part Added##############################Start
        driver.close()
        print("line 777")
    
        pattern = r'\d{2}-\d{2}-\d{4}'
        match = re.search(pattern, aq)
        Splitting = (match.group())
        Replaced = Splitting.replace('-', '/')
        print(Replaced)
    
        print("line 785")
        Effective_Date = datetime.strptime(Replaced, '%m/%d/%Y')
        print(Effective_Date)
        print("line 788")
    
        STATE_LEIN_DATE = Effective_Date - timedelta(days=365 * 20)  # Qualifying Date For Sate Tax
        FED_TAX_DATE = Effective_Date - timedelta(days=365 * 10)  # Qualifying Date For Fed Tax
        UCC_DATE = Effective_Date - timedelta(days=365 * 5)  # Qualifying Date For UCC
        JDG_DATE = Effective_Date - timedelta(days=365 * 20)  # Qualifying Date For Judgemnt
        HOA_DATE = Effective_Date - timedelta(days=365 * 5)  ##Qualifying Date For HOA
        print("line 795")
        time.sleep(2)
        df['Doc Recorded'] = pd.to_datetime(df['Doc Recorded'], format='%m/%d/%Y')
        print("line 798")
        Liens = df['Doc Type'].str.contains('JUDGMENT|LIEN|STATE LIEN|FEDERAL LIEN|UCC|HOA',
                                            case=False, regex=True)
        filterd_df = df[Liens]
        print("line 802")
        FED = filterd_df[
            ((filterd_df['Doc Type'] == 'FEDERAL LIEN') & (filterd_df['Doc Recorded'] >= FED_TAX_DATE))]
        ST = filterd_df[
            ((filterd_df['Doc Type'] == 'STATE LIEN') & (filterd_df['Doc Recorded'] >= STATE_LEIN_DATE))]
        UCC = filterd_df[((filterd_df['Doc Type'] == 'UCC') & (filterd_df['Doc Recorded'] >= UCC_DATE))]
        JUD = filterd_df[
            ((filterd_df['Doc Type'] == 'JUDGMENT') & (filterd_df['Doc Recorded'] >= JDG_DATE))]
        HOA = filterd_df[((filterd_df['Doc Type'] == 'HOA') & (filterd_df['Doc Recorded'] >= HOA_DATE))]
        print("line 811")
        with pd.ExcelWriter(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                EXCELORDERNO) + '\\LeinDocFN.xlsx') as writer:
            FED.to_excel(writer, sheet_name='Sheet1', index=False, startrow=0)
            ST.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(FED) + 1, header=False)
            UCC.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(ST) + 1, header=False)
            JUD.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(UCC) + 1, header=False)
            HOA.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(JUD) + 1, header=False)
    
        print("line 820")
        dataframe2 = pd.read_excel(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
            EXCELORDERNO) + '\\LeinDocFN.xlsx')
    
        chrome_driver = 'chromedriver_win32\chromedriver.exe'
        time.sleep(1)
        chrome_options = Options()
        chrome_options.add_argument('--kiosk-printing')
        chrome_options.add_argument('--disable-extensions')
        time.sleep(2)
        driver = webdriver.Chrome(options=chrome_options)
        time.sleep(2)
        print("line 833")
        for Document_Number in (dataframe2['Doc Number']):
            time.sleep(1)
            driver.get('https://crs.cookcountyclerkil.gov/Search')
            driver.find_element(By.XPATH,
                                '/html/body/div[2]/div/div[3]/div/div/form/div[2]/div[2]/div[3]/div/div[2]/input').send_keys(
                Document_Number, Keys.ENTER)
            driver.maximize_window()
            time.sleep(5)
            h = driver.find_element(By.XPATH, '//*[@id="tblData"]/tbody/tr/td[2]/a').get_attribute('href')
            driver.get(h)
    
            WebDriverWait(driver, 15).until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a')))
            LienType = driver.find_element(By.XPATH, '//*[@id="divcol1"]/div[1]/table/tbody/tr[2]/td').text
            LienNumber = driver.find_element(By.XPATH,
                                             '//*[@id="divcol1"]/div[1]/table/tbody/tr[1]/td').text
            LinkF = driver.find_element(By.XPATH,
                                        '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a').get_attribute(
                'href')
    
            max_retry = 5
            retry_count = 0
            while retry_count < max_retry:  # to download the pdf Document file
                try:
                    time.sleep(2)
                    r = requests.get(LinkF)
                    time.sleep(1)
    
                    with open(
                            os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                                EXCELORDERNO) + '\\' + LienType + ' ' + LienNumber + '.pdf',
                            'wb') as fd:
                        for chunk in r.iter_content(chunk_size=40):
                            fd.write(chunk)
    
                    break
    
                except Exception as e:
                    retry_count += 1
            #####################Lien Part Added##############################End
    
    
    
    else:
    
        FirstNameExtraction = driver.find_element(By.XPATH, '//table')
    
        df1 = pd.read_html(FirstNameExtraction.get_attribute('outerHTML'))[0]
    
        df1.to_excel(
            os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                EXCELORDERNO) + '\\DataExtraction.xlsx',
            index=False)  # to save First Name Tabular Data
        # print("Line 877")
        driver.execute_script('window.print();')
        time.sleep(5)
        path = os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(EXCELORDERNO)
        name = "Full_Name_search_Index"
        pyautogui.FAILSAFE = False
        pyautogui.typewrite(path + '\\' + name + '.pdf')
        time.sleep(2)
        pyautogui.press('enter')
        time.sleep(6)
    
        #####################Lien Part Added##############################Start
        driver.close()
        # print("line 900")
    
        pattern = r'\d{2}-\d{2}-\d{4}'
        match = re.search(pattern, aq)
        Splitting = (match.group())
        Replaced = Splitting.replace('-', '/')
        # print(Replaced)
    
        # print("line 908")
        Effective_Date = datetime.strptime(Replaced, '%m/%d/%Y')
        # print(Effective_Date)
        # print("line 911")
    
        STATE_LEIN_DATE = Effective_Date - timedelta(days=365 * 20)  # Qualifying Date For Sate Tax
        FED_TAX_DATE = Effective_Date - timedelta(days=365 * 10)  # Qualifying Date For Fed Tax
        UCC_DATE = Effective_Date - timedelta(days=365 * 5)  # Qualifying Date For UCC
        JDG_DATE = Effective_Date - timedelta(days=365 * 20)  # Qualifying Date For Judgemnt
        HOA_DATE = Effective_Date - timedelta(days=365 * 5)  ##Qualifying Date For HOA
        # print("line 918")
    
        df1['Doc Recorded'] = pd.to_datetime(df1['Doc Recorded'], format='%m/%d/%Y')
        # print("line 921")
        Liens = df1['Doc Type'].str.contains('JUDGMENT|LIEN|STATE LIEN|FEDERAL LIEN|UCC|HOA',
                                             case=False, regex=True)
        filterd_df = df1[Liens]
        # print("line 925")
        FED = filterd_df[
            ((filterd_df['Doc Type'] == 'FEDERAL LIEN') & (filterd_df['Doc Recorded'] >= FED_TAX_DATE))]
        ST = filterd_df[
            ((filterd_df['Doc Type'] == 'STATE LIEN') & (filterd_df['Doc Recorded'] >= STATE_LEIN_DATE))]
        UCC = filterd_df[((filterd_df['Doc Type'] == 'UCC') & (filterd_df['Doc Recorded'] >= UCC_DATE))]
        JUD = filterd_df[
            ((filterd_df['Doc Type'] == 'JUDGMENT') & (filterd_df['Doc Recorded'] >= JDG_DATE))]
    HOA = filterd_df[((filterd_df['Doc Type'] == 'HOA') & (filterd_df['Doc Recorded'] >= HOA_DATE))]
    # print("line 934")
    with pd.ExcelWriter(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
            EXCELORDERNO) + '\\LeinDocFN.xlsx') as writer:
        FED.to_excel(writer, sheet_name='Sheet1', index=False, startrow=0)
        ST.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(FED) + 1, header=False)
        UCC.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(ST) + 1, header=False)
        JUD.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(UCC) + 1, header=False)
        HOA.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(JUD) + 1, header=False)
    
    # print("line 943")
    dataframe2 = pd.read_excel(os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
        EXCELORDERNO) + '\\LeinDocFN.xlsx')
    
    chrome_driver = 'chromedriver_win32\chromedriver.exe'
    time.sleep(1)
    chrome_options = Options()
    chrome_options.add_argument('--kiosk-printing')
    chrome_options.add_argument('--disable-extensions')
    time.sleep(2)
    driver = webdriver.Chrome(options=chrome_options)
    time.sleep(2)
    # print("line 955")
    for Document_Number in (dataframe2['Doc Number']):
        time.sleep(1)
        driver.get('https://crs.cookcountyclerkil.gov/Search')
        driver.find_element(By.XPATH,
                            '/html/body/div[2]/div/div[3]/div/div/form/div[2]/div[2]/div[3]/div/div[2]/input').send_keys(
            Document_Number, Keys.ENTER)
        driver.maximize_window()
        time.sleep(5)
        h = driver.find_element(By.XPATH, '//*[@id="tblData"]/tbody/tr/td[2]/a').get_attribute('href')
        driver.get(h)
    
        WebDriverWait(driver, 15).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a')))
        LienType = driver.find_element(By.XPATH, '//*[@id="divcol1"]/div[1]/table/tbody/tr[2]/td').text
        LienNumber = driver.find_element(By.XPATH,
                                         '//*[@id="divcol1"]/div[1]/table/tbody/tr[1]/td').text
        LinkF = driver.find_element(By.XPATH,
                                    '/html/body/div[2]/div/div[3]/div/div/fieldset/div[1]/div[2]/div/div/div/a').get_attribute(
            'href')
    
        max_retry = 5
        retry_count = 0
        while retry_count < max_retry:  # to download the pdf Document file
            try:
                time.sleep(2)
                r = requests.get(LinkF)
                time.sleep(1)
    
                with open(
                        os.getcwd() + "\\Output\\COOK_COUNTY\\" + "Order No " + str(
                            EXCELORDERNO) + '\\' + LienType + ' ' + LienNumber + '.pdf',
                        'wb') as fd:
                    for chunk in r.iter_content(chunk_size=40):
                        fd.write(chunk)
    
                break
    
            except Exception as e:
                retry_count += 1
    #####################Lien Part Added##############################End
    
    
    time.sleep(2)
    
